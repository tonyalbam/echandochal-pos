import sqlite3
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from app.database.schema import create_database
from app.services.cash_closing_service import CashClosingService
from app.services.dashboard_service import DashboardService
from app.services.report_service import ReportService
from app.services.sale_history_service import SaleHistoryService
from app.services.sale_service import SaleService


class MemoryDatabase:
    def __init__(self) -> None:
        self.connection = sqlite3.connect(":memory:")
        self.connection.row_factory = sqlite3.Row
        self.connection.execute("PRAGMA foreign_keys = ON")

    def cursor(self):
        return self.connection.cursor()

    def commit(self) -> None:
        self.connection.commit()

    def rollback(self) -> None:
        self.connection.rollback()

    def close(self) -> None:
        self.connection.close()


class FixedDateTime:
    @classmethod
    def now(cls):
        from datetime import datetime
        return datetime(2026, 8, 31, 12, 0, 0)


class MixedPaymentSaleTest(unittest.TestCase):
    def setUp(self) -> None:
        self.database = MemoryDatabase()
        create_database(self.database)
        cursor = self.database.cursor()
        self.product_ids = []
        for number, price, cost in ((1, 100, 50), (2, 200, 80), (3, 300, 120)):
            cursor.execute(
                """
                INSERT INTO productos (
                    codigo, nombre, costo, precio, existencia
                ) VALUES (?, ?, ?, ?, ?)
                """,
                (f"P-{number}", f"Producto {number}", cost, price, 10),
            )
            self.product_ids.append(int(cursor.lastrowid))
        self.database.commit()

    def tearDown(self) -> None:
        self.database.close()

    def _create_sale(self) -> dict:
        items = [
            {
                "producto_id": self.product_ids[0], "cantidad": 1,
                "precio_unitario": 100, "metodo_pago": "Efectivo",
            },
            {
                "producto_id": self.product_ids[1], "cantidad": 1,
                "precio_unitario": 200, "metodo_pago": "Transferencia",
            },
            {
                "producto_id": self.product_ids[2], "cantidad": 1,
                "precio_unitario": 300, "metodo_pago": "Mercado Libre",
            },
        ]
        with patch("app.services.sale_service.datetime", FixedDateTime):
            return SaleService(self.database).create_sale(items, discount=60)

    def test_sale_registers_payment_and_commission_per_item(self) -> None:
        sale = self._create_sale()
        self.assertEqual(sale["metodo_pago"], "Mixto")
        self.assertEqual(sale["total"], 540)
        self.assertEqual(sale["monto_comision"], 10.8)
        self.assertEqual(sale["total_neto"], 529.2)

        detail = SaleHistoryService(self.database).get_sale(sale["id"])
        self.assertEqual(
            [item["metodo_pago"] for item in detail["items"]],
            ["Efectivo", "Transferencia", "Mercado Libre"],
        )
        self.assertEqual(
            [item["descuento"] for item in detail["items"]],
            [10, 20, 30],
        )
        self.assertEqual(detail["items"][2]["monto_comision"], 10.8)

    def test_dashboard_reports_and_cash_closing_use_item_payments(self) -> None:
        self._create_sale()
        dashboard = DashboardService(self.database)
        payments = dashboard.get_payment_sales_for_period("2026-08")
        self.assertEqual(payments["Efectivo"], 90)
        self.assertEqual(payments["Transferencia"], 180)
        self.assertEqual(payments["Mercado Libre"], 270)

        report = ReportService(self.database).get_annual_financial_report(2026)
        self.assertEqual(report["ventas"], 540)
        self.assertEqual(report["comisiones"], 10.8)
        self.assertEqual(report["utilidad"], 279.2)
        self.assertEqual(report["formas_pago"][2]["ventas"], 270)

        closing_service = CashClosingService(self.database)
        closing = closing_service.get_daily_closing("2026-08-31")
        self.assertEqual(closing["formas_pago"][0]["ventas"], 90)
        self.assertEqual(closing["formas_pago"][1]["ventas"], 180)
        self.assertEqual(closing["formas_pago"][2]["ventas"], 270)

        cash = closing_service.calculate_cash_reconciliation(
            "2026-08-31",
            {("Billete", 50.0): 1, ("Billete", 20.0): 2},
        )
        self.assertEqual(cash["efectivo_esperado"], 90)
        self.assertEqual(cash["efectivo_contado"], 90)
        self.assertEqual(cash["diferencia"], 0)
        self.assertEqual(cash["estado"], "CUADRA")

    def test_exported_reports_show_item_payment_breakdown(self) -> None:
        self._create_sale()
        with tempfile.TemporaryDirectory() as directory:
            history_path = SaleHistoryService(self.database).export_sales_report(
                Path(directory) / "historial.xlsx"
            )
            annual_path = ReportService(self.database).export_annual_financial_report(
                2026, Path(directory) / "financiero.xlsx"
            )
            history = load_workbook(history_path, data_only=False)
            items = history["Productos vendidos"]
            self.assertEqual(items["J2"].value, "Efectivo")
            self.assertEqual(items["J4"].value, "Mercado Libre")
            self.assertEqual(items["L4"].value, 10.8)
            history.close()

            annual = load_workbook(annual_path, data_only=False)
            payments = annual["Formas de pago"]
            self.assertEqual(payments["A4"].value, "Mercado Libre")
            self.assertEqual(payments["B4"].value, 270)
            annual.close()


if __name__ == "__main__":
    unittest.main()
