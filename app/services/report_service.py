from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.database.connection import Database
from app.services.configuration_service import ConfigurationService


class ReportService:
    """Consultas y cálculos para reportes financieros."""
    PAYMENT_METHODS = ("Efectivo", "Transferencia", "Mercado Libre")

    def __init__(self, database: Database) -> None:
        self.database = database

    def get_annual_financial_report(
        self,
        year: int,
    ) -> dict:
        """
        Devuelve el resumen financiero anual y
        el detalle mensual del año indicado.
        """

        cursor = self.database.cursor()

        cursor.execute(
            """
            SELECT
                substr(v.fecha, 6, 2) AS mes,
                COALESCE(SUM(
                    CASE
                        WHEN dv.metodo_pago IS NULL AND v.subtotal > 0
                        THEN v.total * dv.subtotal / v.subtotal
                        ELSE dv.subtotal - dv.descuento
                    END
                ), 0) AS ventas,
                COALESCE(SUM(
                    COALESCE(
                        dv.monto_comision,
                        CASE WHEN v.subtotal > 0
                            THEN v.monto_comision * dv.subtotal / v.subtotal
                            ELSE 0 END
                    )
                ), 0) AS comisiones,
                COALESCE(SUM(dv.cantidad * dv.costo_unitario), 0) AS costo
            FROM ventas v
            INNER JOIN detalle_venta dv ON dv.venta_id = v.id
            WHERE v.cancelada = 0
              AND substr(v.fecha, 1, 4) = ?
            GROUP BY substr(v.fecha, 6, 2)
            ORDER BY mes
            """,
            (str(year),),
        )

        rows = cursor.fetchall()

        cursor.execute(
            """
            SELECT
                COALESCE(dv.metodo_pago, v.metodo_pago) AS metodo_pago,
                COALESCE(SUM(
                    CASE
                        WHEN dv.metodo_pago IS NULL AND v.subtotal > 0
                        THEN v.total * dv.subtotal / v.subtotal
                        ELSE dv.subtotal - dv.descuento
                    END
                ), 0) AS ventas,
                COALESCE(SUM(
                    COALESCE(
                        dv.monto_comision,
                        CASE WHEN v.subtotal > 0
                            THEN v.monto_comision * dv.subtotal / v.subtotal
                            ELSE 0 END
                    )
                ), 0) AS comisiones
            FROM detalle_venta dv
            INNER JOIN ventas v ON v.id = dv.venta_id
            WHERE v.cancelada = 0
              AND substr(v.fecha, 1, 4) = ?
            GROUP BY COALESCE(dv.metodo_pago, v.metodo_pago)
            """,
            (str(year),),
        )
        payment_rows = {
            row["metodo_pago"]: dict(row)
            for row in cursor.fetchall()
        }
        payments = []
        for method in self.PAYMENT_METHODS:
            values = payment_rows.get(method, {})
            sales = round(float(values.get("ventas", 0)), 2)
            commissions = round(float(values.get("comisiones", 0)), 2)
            payments.append({
                "metodo_pago": method,
                "ventas": sales,
                "comisiones": commissions,
                "ingreso_neto": round(sales - commissions, 2),
            })

        data_by_month = {
            int(row["mes"]): {
                "ventas": float(row["ventas"]),
                "comisiones": float(
                    row["comisiones"]
                ),
                "costo": float(row["costo"]),
            }
            for row in rows
        }

        monthly = []

        total_sales = 0.0
        total_commissions = 0.0
        total_cost = 0.0

        for month in range(1, 13):
            values = data_by_month.get(
                month,
                {
                    "ventas": 0.0,
                    "comisiones": 0.0,
                    "costo": 0.0,
                },
            )

            sales = values["ventas"]
            commissions = values["comisiones"]
            cost = values["costo"]

            net_income = sales - commissions
            profit = sales - cost - commissions

            monthly.append(
                {
                    "mes": month,
                    "ventas": round(sales, 2),
                    "comisiones": round(
                        commissions,
                        2,
                    ),
                    "costo": round(cost, 2),
                    "ingreso_neto": round(
                        net_income,
                        2,
                    ),
                    "utilidad": round(
                        profit,
                        2,
                    ),
                }
            )

            total_sales += sales
            total_commissions += commissions
            total_cost += cost

        total_net_income = (
            total_sales - total_commissions
        )

        total_profit = (
            total_sales
            - total_cost
            - total_commissions
        )

        return {
            "year": year,
            "ventas": round(total_sales, 2),
            "costo": round(total_cost, 2),
            "comisiones": round(
                total_commissions,
                2,
            ),
            "ingreso_neto": round(
                total_net_income,
                2,
            ),
            "utilidad": round(
                total_profit,
                2,
            ),
            "mensual": monthly,
            "formas_pago": payments,
        }

    def export_annual_financial_report(
        self,
        year: int,
        destination: str | Path,
    ) -> Path:
        """Exporta el reporte financiero anual a un archivo Excel."""

        report = self.get_annual_financial_report(year)
        output_path = Path(destination)

        if output_path.suffix.lower() != ".xlsx":
            output_path = output_path.with_suffix(".xlsx")

        output_path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Reporte anual"
        payments_sheet = workbook.create_sheet("Formas de pago")

        title_fill = PatternFill("solid", fgColor="1F4E78")
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        title_font = Font(color="FFFFFF", bold=True, size=16)
        header_font = Font(bold=True)
        money_format = '"$"#,##0.00'

        sheet.merge_cells("A1:F1")
        business_name = ConfigurationService(
            self.database
        ).get_business_name()
        sheet["A1"] = f"{business_name} POS - Reporte financiero {year}"
        sheet["A1"].fill = title_fill
        sheet["A1"].font = title_font
        sheet["A1"].alignment = Alignment(horizontal="center")

        summary = (
            ("Ventas", report["ventas"]),
            ("Costo", report["costo"]),
            ("Comisiones", report["comisiones"]),
            ("Ingreso neto", report["ingreso_neto"]),
            ("Utilidad", report["utilidad"]),
        )

        for column, (label, value) in enumerate(summary, start=1):
            label_cell = sheet.cell(row=3, column=column, value=label)
            value_cell = sheet.cell(row=4, column=column, value=value)
            label_cell.fill = header_fill
            label_cell.font = header_font
            label_cell.alignment = Alignment(horizontal="center")
            value_cell.number_format = money_format
            value_cell.alignment = Alignment(horizontal="center")

        headers = (
            "Mes",
            "Ventas",
            "Costo",
            "Comisiones",
            "Ingreso neto",
            "Utilidad",
        )

        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(row=6, column=column, value=header)
            cell.fill = title_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        month_names = (
            "Enero", "Febrero", "Marzo", "Abril",
            "Mayo", "Junio", "Julio", "Agosto",
            "Septiembre", "Octubre", "Noviembre", "Diciembre",
        )

        for row_number, month in enumerate(report["mensual"], start=7):
            values = (
                month_names[month["mes"] - 1],
                month["ventas"],
                month["costo"],
                month["comisiones"],
                month["ingreso_neto"],
                month["utilidad"],
            )

            for column, value in enumerate(values, start=1):
                cell = sheet.cell(row=row_number, column=column, value=value)
                if column > 1:
                    cell.number_format = money_format

        total_row = 19
        sheet.cell(row=total_row, column=1, value="Total anual")
        sheet.cell(row=total_row, column=1).font = header_font

        total_keys = (
            "ventas", "costo", "comisiones", "ingreso_neto", "utilidad"
        )
        for column, key in enumerate(total_keys, start=2):
            cell = sheet.cell(
                row=total_row,
                column=column,
                value=report[key],
            )
            cell.font = header_font
            cell.number_format = money_format

        sheet.freeze_panes = "A7"
        sheet.auto_filter.ref = "A6:F18"

        widths = (16, 15, 15, 15, 17, 15)
        for column, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(column)].width = width

        payment_headers = (
            "Forma de pago", "Ventas", "Comisiones", "Ingreso neto"
        )
        for column, header in enumerate(payment_headers, start=1):
            cell = payments_sheet.cell(row=1, column=column, value=header)
            cell.fill = title_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        for row_number, payment in enumerate(report["formas_pago"], start=2):
            values = (
                payment["metodo_pago"], payment["ventas"],
                payment["comisiones"], payment["ingreso_neto"],
            )
            for column, value in enumerate(values, start=1):
                cell = payments_sheet.cell(row=row_number, column=column, value=value)
                if column > 1:
                    cell.number_format = money_format
        payments_sheet.freeze_panes = "A2"
        payments_sheet.auto_filter.ref = "A1:D4"
        for column, width in enumerate((22, 16, 16, 17), start=1):
            payments_sheet.column_dimensions[get_column_letter(column)].width = width

        workbook.save(output_path)
        return output_path
