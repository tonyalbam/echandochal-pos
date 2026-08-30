from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.database.connection import Database


class SaleHistoryService:
    """Consulta el historial, detalle y cancelación de ventas."""

    def __init__(self, database: Database) -> None:
        self.database = database

    def list_sales(
        self,
        search: str = "",
        include_cancelled: bool = True,
        date_from: str | None = None,
        date_to: str | None = None,
    ) -> list[dict]:

        cursor = self.database.cursor()

        sql = """
            SELECT
                v.id,
                v.folio,
                v.fecha,
                v.hora,
                v.subtotal,
                v.descuento,
                v.total,
                v.metodo_pago,
                v.porcentaje_comision,
                v.monto_comision,
                v.total_neto,
                v.cancelada
            FROM ventas v
            WHERE 1 = 1
        """

        parameters: list = []

        if not include_cancelled:
            sql += " AND v.cancelada = 0"

        if date_from:
            sql += " AND v.fecha >= ?"
            parameters.append(date_from)

        if date_to:
            sql += " AND v.fecha <= ?"
            parameters.append(date_to)

        search = search.strip()

        if search:
            sql += """
                AND (
                    v.folio LIKE ?
                    OR v.metodo_pago LIKE ?
                )
            """

            pattern = f"%{search}%"
            parameters.extend([pattern, pattern])

        sql += """
            ORDER BY
                v.fecha DESC,
                v.hora DESC,
                v.id DESC
        """

        cursor.execute(sql, parameters)

        return [
            dict(row)
            for row in cursor.fetchall()
        ]

    def get_sale(
        self,
        sale_id: int,
    ) -> dict | None:

        cursor = self.database.cursor()

        cursor.execute(
            """
            SELECT
                v.id,
                v.folio,
                v.fecha,
                v.hora,
                v.subtotal,
                v.descuento,
                v.total,
                v.metodo_pago,
                v.porcentaje_comision,
                v.monto_comision,
                v.total_neto,
                v.usuario_id,
                v.cancelada
            FROM ventas v
            WHERE v.id = ?
            """,
            (sale_id,),
        )

        row = cursor.fetchone()

        if row is None:
            return None

        sale = dict(row)

        cursor.execute(
            """
            SELECT
                dv.id,
                dv.producto_id,
                p.codigo,
                p.nombre,
                dv.cantidad,
                dv.precio_unitario,
                dv.costo_unitario,
                dv.descuento,
                dv.subtotal
            FROM detalle_venta dv
            INNER JOIN productos p
                ON p.id = dv.producto_id
            WHERE dv.venta_id = ?
            ORDER BY dv.id
            """,
            (sale_id,),
        )

        sale["items"] = [
            dict(item)
            for item in cursor.fetchall()
        ]

        return sale

    def export_sales_report(
        self,
        destination: str | Path,
        search: str = "",
        include_cancelled: bool = True,
        date_from: str | None = None,
        date_to: str | None = None,
    ) -> Path:
        """Exporta las ventas filtradas y sus productos a Excel."""

        if date_from and date_to and date_from > date_to:
            raise ValueError(
                "La fecha inicial no puede ser posterior a la fecha final."
            )

        sales = self.list_sales(
            search=search,
            include_cancelled=include_cancelled,
            date_from=date_from,
            date_to=date_to,
        )

        output_path = Path(destination)
        if output_path.suffix.lower() != ".xlsx":
            output_path = output_path.with_suffix(".xlsx")
        output_path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        sales_sheet = workbook.active
        sales_sheet.title = "Ventas"
        items_sheet = workbook.create_sheet("Productos vendidos")

        title_fill = PatternFill("solid", fgColor="1F4E78")
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        white_bold_font = Font(color="FFFFFF", bold=True)
        bold_font = Font(bold=True)
        money_format = '"$"#,##0.00'

        period = "Todas las fechas"
        if date_from or date_to:
            period = f"{date_from or 'Inicio'} a {date_to or 'Hoy'}"

        sales_sheet.merge_cells("A1:L1")
        sales_sheet["A1"] = "Echando Chal POS - Detalle de ventas"
        sales_sheet["A1"].fill = title_fill
        sales_sheet["A1"].font = Font(
            color="FFFFFF", bold=True, size=16
        )
        sales_sheet["A1"].alignment = Alignment(horizontal="center")
        sales_sheet["A2"] = f"Periodo: {period}"
        sales_sheet["A3"] = f"Generado: {date.today().isoformat()}"

        headers = (
            "Folio", "Fecha", "Hora", "Forma de pago", "Subtotal",
            "Descuento", "Total", "% comisión", "Comisión",
            "Ingreso neto", "Estado", "Productos",
        )
        for column, header in enumerate(headers, start=1):
            cell = sales_sheet.cell(row=5, column=column, value=header)
            cell.fill = title_fill
            cell.font = white_bold_font
            cell.alignment = Alignment(horizontal="center")

        item_headers = (
            "Folio", "Fecha", "Código", "Producto", "Cantidad",
            "Precio unitario", "Costo unitario", "Importe", "Estado",
        )
        for column, header in enumerate(item_headers, start=1):
            cell = items_sheet.cell(row=1, column=column, value=header)
            cell.fill = title_fill
            cell.font = white_bold_font
            cell.alignment = Alignment(horizontal="center")

        item_row = 2
        for sales_row, sale in enumerate(sales, start=6):
            detail = self.get_sale(int(sale["id"]))
            items = detail["items"] if detail else []
            status = "CANCELADA" if sale["cancelada"] else "ACTIVA"
            values = (
                sale["folio"], sale["fecha"], sale["hora"],
                sale["metodo_pago"], sale["subtotal"], sale["descuento"],
                sale["total"], sale["porcentaje_comision"] / 100,
                sale["monto_comision"], sale["total_neto"], status,
                len(items),
            )
            for column, value in enumerate(values, start=1):
                cell = sales_sheet.cell(
                    row=sales_row, column=column, value=value
                )
                if column in (5, 6, 7, 9, 10):
                    cell.number_format = money_format
                elif column == 8:
                    cell.number_format = "0.00%"

            for item in items:
                item_values = (
                    sale["folio"], sale["fecha"], item["codigo"],
                    item["nombre"], item["cantidad"],
                    item["precio_unitario"], item["costo_unitario"],
                    item["subtotal"], status,
                )
                for column, value in enumerate(item_values, start=1):
                    cell = items_sheet.cell(
                        row=item_row, column=column, value=value
                    )
                    if column in (6, 7, 8):
                        cell.number_format = money_format
                item_row += 1

        total_row = max(len(sales) + 6, 7)
        sales_sheet.cell(row=total_row, column=4, value="Totales")
        sales_sheet.cell(row=total_row, column=4).font = bold_font
        for column in (5, 6, 7, 9, 10):
            cell = sales_sheet.cell(
                row=total_row,
                column=column,
                value=f"=SUM({get_column_letter(column)}6:{get_column_letter(column)}{total_row - 1})",
            )
            cell.font = bold_font
            cell.fill = header_fill
            cell.number_format = money_format

        sales_sheet.freeze_panes = "A6"
        if sales:
            sales_sheet.auto_filter.ref = f"A5:L{total_row - 1}"
        items_sheet.freeze_panes = "A2"
        if item_row > 2:
            items_sheet.auto_filter.ref = f"A1:I{item_row - 1}"

        sales_widths = (22, 13, 11, 19, 14, 14, 14, 13, 14, 15, 13, 11)
        item_widths = (22, 13, 15, 34, 12, 16, 16, 14, 13)
        for column, width in enumerate(sales_widths, start=1):
            sales_sheet.column_dimensions[get_column_letter(column)].width = width
        for column, width in enumerate(item_widths, start=1):
            items_sheet.column_dimensions[get_column_letter(column)].width = width

        workbook.save(output_path)
        return output_path

    def cancel_sale(
        self,
        sale_id: int,
        usuario_id: int | None = None,
    ) -> dict:
        """Cancela una venta y devuelve sus productos al inventario."""

        cursor = self.database.cursor()

        try:
            cursor.execute(
                """
                SELECT
                    id,
                    folio,
                    cancelada
                FROM ventas
                WHERE id = ?
                """,
                (sale_id,),
            )

            venta = cursor.fetchone()

            if venta is None:
                raise ValueError(
                    "La venta no existe."
                )

            if venta["cancelada"]:
                raise ValueError(
                    "La venta ya está cancelada."
                )

            folio = venta["folio"]

            cursor.execute(
                """
                SELECT
                    dv.producto_id,
                    dv.cantidad,
                    p.nombre,
                    p.existencia
                FROM detalle_venta dv
                INNER JOIN productos p
                    ON p.id = dv.producto_id
                WHERE dv.venta_id = ?
                ORDER BY dv.id
                """,
                (sale_id,),
            )

            detalles = cursor.fetchall()

            if not detalles:
                raise ValueError(
                    "La venta no tiene productos asociados."
                )

            for detalle in detalles:
                producto_id = detalle["producto_id"]
                cantidad = float(detalle["cantidad"])

                existencia_anterior = float(
                    detalle["existencia"]
                )

                existencia_nueva = (
                    existencia_anterior + cantidad
                )

                cursor.execute(
                    """
                    UPDATE productos
                    SET
                        existencia = ?,
                        fecha_actualizacion = CURRENT_TIMESTAMP
                    WHERE id = ?
                    """,
                    (
                        existencia_nueva,
                        producto_id,
                    ),
                )

                cursor.execute(
                    """
                    INSERT INTO movimientos_inventario (
                        producto_id,
                        tipo,
                        cantidad,
                        existencia_anterior,
                        existencia_nueva,
                        referencia,
                        usuario_id
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        producto_id,
                        "DEVOLUCION_VENTA",
                        cantidad,
                        existencia_anterior,
                        existencia_nueva,
                        folio,
                        usuario_id,
                    ),
                )

            cursor.execute(
                """
                UPDATE ventas
                SET
                    cancelada = 1
                WHERE id = ?
                """,
                (sale_id,),
            )

            cursor.execute(
                """
                INSERT INTO auditoria (
                    usuario_id,
                    accion,
                    modulo,
                    referencia,
                    detalles
                )
                VALUES (?, ?, ?, ?, ?)
                """,
                (
                    usuario_id,
                    "CANCELAR",
                    "VENTAS",
                    folio,
                    "Venta cancelada y mercancía devuelta al inventario.",
                ),
            )

            self.database.commit()

            return {
                "id": sale_id,
                "folio": folio,
                "productos_devueltos": len(detalles),
            }

        except Exception:
            self.database.rollback()
            raise
