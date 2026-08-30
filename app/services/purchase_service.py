from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.database.connection import Database
from app.services.configuration_service import ConfigurationService


class PurchaseService:
    """Operaciones de negocio relacionadas con compras."""

    def __init__(self, database: Database) -> None:
        self.database = database

    def create_purchase(
        self,
        items: list[dict],
        supplier_id: Optional[int] = None,
        notes: str = "",
        user_id: Optional[int] = None,
    ) -> dict:
        """
        Registra una compra completa.

        items debe contener:
            producto_id
            cantidad
            costo_unitario
        """

        if not items:
            raise ValueError(
                "La compra no contiene productos."
            )

        subtotal = 0.0

        for item in items:
            quantity = float(item["cantidad"])
            unit_cost = float(item["costo_unitario"])

            if quantity <= 0:
                raise ValueError(
                    "La cantidad debe ser mayor que cero."
                )

            if unit_cost < 0:
                raise ValueError(
                    "El costo no puede ser negativo."
                )

            subtotal += quantity * unit_cost

        subtotal = round(subtotal, 2)
        total = subtotal

        now = datetime.now()

        fecha = now.strftime("%Y-%m-%d")
        folio = self._next_folio(fecha)

        cursor = self.database.cursor()

        try:
            cursor.execute("BEGIN")

            # Validamos todos los productos antes de modificar datos.
            for item in items:
                product_id = int(item["producto_id"])

                cursor.execute(
                    """
                    SELECT id, nombre, existencia, activo
                    FROM productos
                    WHERE id = ?
                    """,
                    (product_id,),
                )

                product = cursor.fetchone()

                if not product:
                    raise ValueError(
                        f"El producto con ID {product_id} no existe."
                    )

                if not product["activo"]:
                    raise ValueError(
                        f"El producto '{product['nombre']}' está inactivo."
                    )

            # Cabecera de la compra.
            cursor.execute(
                """
                INSERT INTO compras (
                    folio,
                    proveedor_id,
                    fecha,
                    subtotal,
                    total,
                    notas,
                    usuario_id
                )
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    folio,
                    supplier_id,
                    fecha,
                    subtotal,
                    total,
                    notes,
                    user_id,
                ),
            )

            purchase_id = int(cursor.lastrowid)

            # Detalle + actualización del inventario.
            for item in items:
                product_id = int(item["producto_id"])
                quantity = float(item["cantidad"])
                unit_cost = float(item["costo_unitario"])

                line_subtotal = round(
                    quantity * unit_cost,
                    2,
                )

                cursor.execute(
                    """
                    INSERT INTO detalle_compra (
                        compra_id,
                        producto_id,
                        cantidad,
                        costo_unitario,
                        subtotal
                    )
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (
                        purchase_id,
                        product_id,
                        quantity,
                        unit_cost,
                        line_subtotal,
                    ),
                )

                cursor.execute(
                    """
                    SELECT existencia
                    FROM productos
                    WHERE id = ?
                    """,
                    (product_id,),
                )

                previous_stock = float(
                    cursor.fetchone()["existencia"]
                )

                new_stock = previous_stock + quantity

                cursor.execute(
                    """
                    UPDATE productos
                    SET
                        existencia = ?,
                        costo = ?,
                        fecha_actualizacion = CURRENT_TIMESTAMP
                    WHERE id = ?
                    """,
                    (
                        new_stock,
                        unit_cost,
                        product_id,
                    ),
                )

                cursor.execute(
                    """
                    INSERT INTO movimientos_inventario (
                        producto_id,
                        tipo,
                        cantidad,
                        existencia_anterior,
                        existencia_nueva,
                        referencia,
                        usuario_id
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        product_id,
                        "COMPRA",
                        quantity,
                        previous_stock,
                        new_stock,
                        folio,
                        user_id,
                    ),
                )

            self.database.commit()

        except Exception:
            self.database.rollback()
            raise

        return {
            "id": purchase_id,
            "folio": folio,
            "fecha": fecha,
            "subtotal": subtotal,
            "total": total,
            "proveedor_id": supplier_id,
            "notas": notes,
        }

    def get_purchase(
        self,
        purchase_id: int,
    ) -> dict | None:
        """Obtiene una compra con su detalle."""

        cursor = self.database.cursor()

        cursor.execute(
            """
            SELECT
                c.id,
                c.folio,
                c.proveedor_id,
                c.fecha,
                c.subtotal,
                c.total,
                c.notas,
                c.usuario_id,
                COALESCE(pr.nombre, 'Sin proveedor') AS proveedor
            FROM compras c
            LEFT JOIN proveedores pr
                ON pr.id = c.proveedor_id
            WHERE c.id = ?
            """,
            (purchase_id,),
        )

        row = cursor.fetchone()

        if row is None:
            return None

        purchase = dict(row)

        cursor.execute(
            """
            SELECT
                dc.id,
                dc.producto_id,
                p.codigo,
                p.nombre,
                dc.cantidad,
                dc.costo_unitario,
                dc.subtotal
            FROM detalle_compra dc
            INNER JOIN productos p
                ON p.id = dc.producto_id
            WHERE dc.compra_id = ?
            ORDER BY dc.id
            """,
            (purchase_id,),
        )

        purchase["items"] = [
            dict(item)
            for item in cursor.fetchall()
        ]

        return purchase

    def list_purchases(
        self,
        search: str = "",
        supplier_id: int | None = None,
        date_from: str | None = None,
        date_to: str | None = None,
    ) -> list[dict]:
        """Lista las compras registradas."""

        cursor = self.database.cursor()

        sql = """
            SELECT
                c.id,
                c.folio,
                c.proveedor_id,
                c.fecha,
                c.subtotal,
                c.total,
                c.notas,
                c.usuario_id,
                COALESCE(pr.nombre, 'Sin proveedor') AS proveedor
            FROM compras c
            LEFT JOIN proveedores pr
                ON pr.id = c.proveedor_id
            WHERE 1 = 1
        """

        parameters: list = []

        search = search.strip()

        if supplier_id is not None:
            sql += " AND c.proveedor_id = ?"
            parameters.append(supplier_id)

        if date_from:
            sql += " AND c.fecha >= ?"
            parameters.append(date_from)

        if date_to:
            sql += " AND c.fecha <= ?"
            parameters.append(date_to)

        if search:
            sql += """
                AND (
                    c.folio LIKE ?
                    OR COALESCE(pr.nombre, '') LIKE ?
                    OR COALESCE(c.notas, '') LIKE ?
                )
            """

            pattern = f"%{search}%"
            parameters.extend([pattern, pattern, pattern])

        sql += """
            ORDER BY
                c.fecha DESC,
                c.id DESC
        """

        cursor.execute(
            sql,
            parameters,
        )

        return [
            dict(row)
            for row in cursor.fetchall()
        ]

    def export_purchases_report(
        self,
        destination: str | Path,
        search: str = "",
        supplier_id: int | None = None,
        date_from: str | None = None,
        date_to: str | None = None,
    ) -> Path:
        """Exporta las compras filtradas y sus productos a Excel."""

        if date_from and date_to and date_from > date_to:
            raise ValueError(
                "La fecha inicial no puede ser posterior a la fecha final."
            )

        purchases = self.list_purchases(
            search=search,
            supplier_id=supplier_id,
            date_from=date_from,
            date_to=date_to,
        )

        output_path = Path(destination)
        if output_path.suffix.lower() != ".xlsx":
            output_path = output_path.with_suffix(".xlsx")
        output_path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        purchases_sheet = workbook.active
        purchases_sheet.title = "Compras"
        items_sheet = workbook.create_sheet("Productos comprados")

        title_fill = PatternFill("solid", fgColor="1F4E78")
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        white_bold_font = Font(color="FFFFFF", bold=True)
        bold_font = Font(bold=True)
        money_format = '"$"#,##0.00'

        period = "Todas las fechas"
        if date_from or date_to:
            period = f"{date_from or 'Inicio'} a {date_to or 'Hoy'}"

        purchases_sheet.merge_cells("A1:G1")
        business_name = ConfigurationService(
            self.database
        ).get_business_name()
        purchases_sheet["A1"] = f"{business_name} POS - Historial de compras"
        purchases_sheet["A1"].fill = title_fill
        purchases_sheet["A1"].font = Font(
            color="FFFFFF", bold=True, size=16
        )
        purchases_sheet["A1"].alignment = Alignment(horizontal="center")
        purchases_sheet["A2"] = f"Periodo: {period}"
        purchases_sheet["A3"] = f"Compras encontradas: {len(purchases)}"

        headers = (
            "Folio", "Fecha", "Proveedor", "Subtotal", "Total",
            "Productos", "Notas",
        )
        for column, header in enumerate(headers, start=1):
            cell = purchases_sheet.cell(row=5, column=column, value=header)
            cell.fill = title_fill
            cell.font = white_bold_font
            cell.alignment = Alignment(horizontal="center")

        item_headers = (
            "Folio", "Fecha", "Proveedor", "Código", "Producto",
            "Cantidad", "Costo unitario", "Importe",
        )
        for column, header in enumerate(item_headers, start=1):
            cell = items_sheet.cell(row=1, column=column, value=header)
            cell.fill = title_fill
            cell.font = white_bold_font
            cell.alignment = Alignment(horizontal="center")

        item_row = 2
        for purchase_row, purchase in enumerate(purchases, start=6):
            detail = self.get_purchase(int(purchase["id"]))
            items = detail["items"] if detail else []
            values = (
                purchase["folio"], purchase["fecha"], purchase["proveedor"],
                purchase["subtotal"], purchase["total"], len(items),
                purchase["notas"] or "",
            )
            for column, value in enumerate(values, start=1):
                cell = purchases_sheet.cell(
                    row=purchase_row, column=column, value=value
                )
                if column in (4, 5):
                    cell.number_format = money_format

            for item in items:
                item_values = (
                    purchase["folio"], purchase["fecha"],
                    purchase["proveedor"], item["codigo"], item["nombre"],
                    item["cantidad"], item["costo_unitario"],
                    item["subtotal"],
                )
                for column, value in enumerate(item_values, start=1):
                    cell = items_sheet.cell(
                        row=item_row, column=column, value=value
                    )
                    if column in (7, 8):
                        cell.number_format = money_format
                item_row += 1

        total_row = max(len(purchases) + 6, 7)
        purchases_sheet.cell(row=total_row, column=3, value="Totales")
        purchases_sheet.cell(row=total_row, column=3).font = bold_font
        for column in (4, 5):
            letter = get_column_letter(column)
            cell = purchases_sheet.cell(
                row=total_row,
                column=column,
                value=f"=SUM({letter}6:{letter}{total_row - 1})",
            )
            cell.fill = header_fill
            cell.font = bold_font
            cell.number_format = money_format

        purchases_sheet.freeze_panes = "A6"
        if purchases:
            purchases_sheet.auto_filter.ref = f"A5:G{len(purchases) + 5}"
        items_sheet.freeze_panes = "A2"
        if item_row > 2:
            items_sheet.auto_filter.ref = f"A1:H{item_row - 1}"

        purchase_widths = (22, 13, 28, 15, 15, 12, 40)
        item_widths = (22, 13, 28, 16, 34, 12, 16, 15)
        for column, width in enumerate(purchase_widths, start=1):
            purchases_sheet.column_dimensions[get_column_letter(column)].width = width
        for column, width in enumerate(item_widths, start=1):
            items_sheet.column_dimensions[get_column_letter(column)].width = width

        workbook.save(output_path)
        return output_path

    def _next_folio(
        self,
        fecha: str,
    ) -> str:
        """Genera un folio consecutivo para el día."""

        prefix = f"C-{fecha.replace('-', '')}-"

        cursor = self.database.cursor()

        cursor.execute(
            """
            SELECT folio
            FROM compras
            WHERE folio LIKE ?
            ORDER BY id DESC
            LIMIT 1
            """,
            (f"{prefix}%",),
        )

        row = cursor.fetchone()

        if not row:
            sequence = 1
        else:
            try:
                sequence = (
                    int(
                        str(row["folio"]).split("-")[-1]
                    )
                    + 1
                )
            except (ValueError, IndexError):
                sequence = 1

        return f"{prefix}{sequence:04d}"
