from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.database.connection import Database
from app.services.configuration_service import ConfigurationService


class ProductService:
    """Operaciones de negocio relacionadas con productos."""

    def __init__(self, database: Database) -> None:
        self.database = database

    def list_products(self, search: str = "") -> list[dict]:
        cursor = self.database.cursor()

        search = search.strip()

        if search:
            pattern = f"%{search}%"

            cursor.execute(
                """
                SELECT
                    p.id,
                    p.codigo,
                    p.codigo_barras,
                    p.nombre,
                    COALESCE(c.nombre, '') AS categoria,
                    COALESCE(p.marca, '') AS marca,
                    COALESCE(p.color, '') AS color,
                    p.unidad,
                    p.costo,
                    p.precio,
                    p.existencia,
                    p.stock_minimo,
                    COALESCE(pr.nombre, '') AS proveedor,
                    p.activo
                FROM productos p
                LEFT JOIN categorias c
                    ON c.id = p.categoria_id
                LEFT JOIN proveedores pr
                    ON pr.id = p.proveedor_id
                WHERE p.activo = 1
                  AND (
                      p.codigo LIKE ?
                      OR COALESCE(p.codigo_barras, '') LIKE ?
                      OR p.nombre LIKE ?
                      OR COALESCE(p.marca, '') LIKE ?
                      OR COALESCE(p.color, '') LIKE ?
                  )
                ORDER BY p.nombre COLLATE NOCASE
                """,
                (pattern, pattern, pattern, pattern, pattern),
            )
        else:
            cursor.execute(
                """
                SELECT
                    p.id,
                    p.codigo,
                    p.codigo_barras,
                    p.nombre,
                    COALESCE(c.nombre, '') AS categoria,
                    COALESCE(p.marca, '') AS marca,
                    COALESCE(p.color, '') AS color,
                    p.unidad,
                    p.costo,
                    p.precio,
                    p.existencia,
                    p.stock_minimo,
                    COALESCE(pr.nombre, '') AS proveedor,
                    p.activo
                FROM productos p
                LEFT JOIN categorias c
                    ON c.id = p.categoria_id
                LEFT JOIN proveedores pr
                    ON pr.id = p.proveedor_id
                WHERE p.activo = 1
                ORDER BY p.nombre COLLATE NOCASE
                """
            )

        return [dict(row) for row in cursor.fetchall()]

    def get_product(self, product_id: int) -> Optional[dict]:
        cursor = self.database.cursor()

        cursor.execute(
            """
            SELECT
                p.*,
                COALESCE(c.nombre, '') AS categoria_nombre,
                COALESCE(pr.nombre, '') AS proveedor_nombre
            FROM productos p
            LEFT JOIN categorias c
                ON c.id = p.categoria_id
            LEFT JOIN proveedores pr
                ON pr.id = p.proveedor_id
            WHERE p.id = ?
            """,
            (product_id,),
        )

        row = cursor.fetchone()

        return dict(row) if row else None

    def create_product(self, data: dict) -> int:
        cursor = self.database.cursor()

        cursor.execute(
            """
            INSERT INTO productos (
                codigo,
                codigo_barras,
                nombre,
                categoria_id,
                marca,
                color,
                unidad,
                costo,
                precio,
                existencia,
                stock_minimo,
                proveedor_id
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                data["codigo"],
                data.get("codigo_barras") or None,
                data["nombre"],
                data.get("categoria_id"),
                data.get("marca") or None,
                data.get("color") or None,
                data.get("unidad") or "pieza",
                data.get("costo", 0),
                data.get("precio", 0),
                data.get("existencia", 0),
                data.get("stock_minimo", 0),
                data.get("proveedor_id"),
            ),
        )

        product_id = cursor.lastrowid

        self.database.commit()

        return int(product_id)

    def update_product(self, product_id: int, data: dict) -> None:
        cursor = self.database.cursor()

        cursor.execute(
            """
            UPDATE productos
            SET
                codigo = ?,
                codigo_barras = ?,
                nombre = ?,
                categoria_id = ?,
                marca = ?,
                color = ?,
                unidad = ?,
                costo = ?,
                precio = ?,
                existencia = ?,
                stock_minimo = ?,
                proveedor_id = ?,
                fecha_actualizacion = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (
                data["codigo"],
                data.get("codigo_barras") or None,
                data["nombre"],
                data.get("categoria_id"),
                data.get("marca") or None,
                data.get("color") or None,
                data.get("unidad") or "pieza",
                data.get("costo", 0),
                data.get("precio", 0),
                data.get("existencia", 0),
                data.get("stock_minimo", 0),
                data.get("proveedor_id"),
                product_id,
            ),
        )

        self.database.commit()

    def deactivate_product(self, product_id: int) -> None:
        cursor = self.database.cursor()

        cursor.execute(
            """
            UPDATE productos
            SET
                activo = 0,
                fecha_actualizacion = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (product_id,),
        )

        self.database.commit()

    def get_categories(self) -> list[dict]:
        cursor = self.database.cursor()

        cursor.execute(
            """
            SELECT id, nombre
            FROM categorias
            WHERE activo = 1
            ORDER BY nombre COLLATE NOCASE
            """
        )

        return [dict(row) for row in cursor.fetchall()]

    def get_suppliers(self) -> list[dict]:
        cursor = self.database.cursor()

        cursor.execute(
            """
            SELECT id, nombre
            FROM proveedores
            WHERE activo = 1
            ORDER BY nombre COLLATE NOCASE
            """
        )

        return [dict(row) for row in cursor.fetchall()]

    def get_inventory_report(self, search: str = "") -> dict:
        """Devuelve el inventario activo y sus indicadores principales."""

        products = self.list_products(search)

        inventory = []
        for product in products:
            stock = float(product["existencia"])
            minimum = float(product["stock_minimo"])
            cost = float(product["costo"])
            price = float(product["precio"])

            if stock <= 0:
                status = "AGOTADO"
            elif stock <= minimum:
                status = "BAJO"
            else:
                status = "OK"

            inventory.append(
                {
                    **product,
                    "faltante_minimo": round(max(minimum - stock, 0), 3),
                    "valor_costo": round(stock * cost, 2),
                    "valor_venta": round(stock * price, 2),
                    "estado_stock": status,
                }
            )

        return {
            "productos": inventory,
            "total_productos": len(inventory),
            "productos_stock_bajo": sum(
                1
                for product in inventory
                if product["estado_stock"] in ("BAJO", "AGOTADO")
            ),
            "valor_total_costo": round(
                sum(product["valor_costo"] for product in inventory),
                2,
            ),
            "valor_total_venta": round(
                sum(product["valor_venta"] for product in inventory),
                2,
            ),
        }

    def export_inventory_report(
        self,
        destination: str | Path,
        search: str = "",
    ) -> Path:
        """Exporta el inventario activo a un archivo Excel."""

        report = self.get_inventory_report(search)
        products = report["productos"]

        output_path = Path(destination)
        if output_path.suffix.lower() != ".xlsx":
            output_path = output_path.with_suffix(".xlsx")
        output_path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Inventario"

        title_fill = PatternFill("solid", fgColor="1F4E78")
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        warning_fill = PatternFill("solid", fgColor="FFF2CC")
        danger_fill = PatternFill("solid", fgColor="F4CCCC")
        white_bold_font = Font(color="FFFFFF", bold=True)
        bold_font = Font(bold=True)
        money_format = '"$"#,##0.00'

        sheet.merge_cells("A1:O1")
        business_name = ConfigurationService(
            self.database
        ).get_business_name()
        sheet["A1"] = f"{business_name} POS - Reporte de inventario"
        sheet["A1"].fill = title_fill
        sheet["A1"].font = Font(color="FFFFFF", bold=True, size=16)
        sheet["A1"].alignment = Alignment(horizontal="center")

        last_data_row = max(len(products) + 6, 7)
        summary = (
            ("Productos activos", f"=COUNTA(A7:A{last_data_row})"),
            ("Valor a costo", f"=SUM(M7:M{last_data_row})"),
            ("Valor potencial", f"=SUM(N7:N{last_data_row})"),
            (
                "Stock bajo/agotado",
                f'=COUNTIF(O7:O{last_data_row},"BAJO")+'
                f'COUNTIF(O7:O{last_data_row},"AGOTADO")',
            ),
        )
        for index, (label, formula) in enumerate(summary):
            label_column = index * 2 + 1
            value_column = label_column + 1
            label_cell = sheet.cell(row=3, column=label_column, value=label)
            value_cell = sheet.cell(row=3, column=value_column, value=formula)
            label_cell.fill = header_fill
            label_cell.font = bold_font
            value_cell.font = bold_font
            if index in (1, 2):
                value_cell.number_format = money_format

        headers = (
            "Código", "Código de barras", "Producto", "Categoría",
            "Marca", "Proveedor", "Unidad", "Costo", "Precio",
            "Existencia", "Stock mínimo", "Faltante al mínimo",
            "Valor a costo", "Valor potencial", "Estado",
        )
        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(row=6, column=column, value=header)
            cell.fill = title_fill
            cell.font = white_bold_font
            cell.alignment = Alignment(horizontal="center")

        for row, product in enumerate(products, start=7):
            base_values = (
                product["codigo"], product["codigo_barras"] or "",
                product["nombre"], product["categoria"], product["marca"],
                product["proveedor"], product["unidad"], product["costo"],
                product["precio"], product["existencia"],
                product["stock_minimo"],
            )
            for column, value in enumerate(base_values, start=1):
                sheet.cell(row=row, column=column, value=value)

            sheet.cell(row=row, column=12, value=f"=MAX(K{row}-J{row},0)")
            sheet.cell(row=row, column=13, value=f"=H{row}*J{row}")
            sheet.cell(row=row, column=14, value=f"=I{row}*J{row}")
            sheet.cell(
                row=row,
                column=15,
                value=(
                    f'=IF(J{row}<=0,"AGOTADO",'
                    f'IF(J{row}<=K{row},"BAJO","OK"))'
                ),
            )

            for column in (8, 9, 13, 14):
                sheet.cell(row=row, column=column).number_format = money_format
            for column in (10, 11, 12):
                sheet.cell(row=row, column=column).number_format = "#,##0.###"

            status = product["estado_stock"]
            if status == "AGOTADO":
                sheet.cell(row=row, column=15).fill = danger_fill
            elif status == "BAJO":
                sheet.cell(row=row, column=15).fill = warning_fill

        sheet.freeze_panes = "A7"
        if products:
            sheet.auto_filter.ref = f"A6:O{len(products) + 6}"

        widths = (
            16, 20, 34, 20, 18, 24, 12, 14, 14, 14, 15, 19, 16, 17, 13
        )
        for column, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(column)].width = width

        workbook.save(output_path)
        return output_path

    def create_category(self, name: str) -> int:
        cursor = self.database.cursor()

        cursor.execute(
            """
            INSERT INTO categorias (nombre)
            VALUES (?)
            """,
            (name.strip(),),
        )

        category_id = cursor.lastrowid

        self.database.commit()

        return int(category_id)

    def create_supplier(self, name: str) -> int:
        cursor = self.database.cursor()

        cursor.execute(
            """
            INSERT INTO proveedores (nombre)
            VALUES (?)
            """,
            (name.strip(),),
        )

        supplier_id = cursor.lastrowid

        self.database.commit()

        return int(supplier_id)
