from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Image,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from app.database.connection import Database
from app.services.configuration_service import ConfigurationService


class CashClosingService:
    """Calcula y exporta el corte diario de ventas."""

    PAYMENT_METHODS = ("Efectivo", "Transferencia", "Mercado Libre")
    DENOMINATIONS = (
        ("Billete", 1000.0), ("Billete", 500.0),
        ("Billete", 200.0), ("Billete", 100.0),
        ("Billete", 50.0), ("Billete", 20.0),
        ("Moneda", 20.0), ("Moneda", 10.0),
        ("Moneda", 5.0), ("Moneda", 2.0),
        ("Moneda", 1.0), ("Moneda", 0.5),
    )

    def __init__(self, database: Database) -> None:
        self.database = database

    def get_daily_closing(self, closing_date: str) -> dict:
        cursor = self.database.cursor()
        cursor.execute(
            """
            SELECT
                COUNT(*) AS tickets,
                COALESCE(SUM(v.total), 0) AS ventas,
                COALESCE(SUM(v.monto_comision), 0) AS comisiones,
                COALESCE(SUM(v.total_neto), 0) AS dinero_neto,
                COALESCE(SUM(costos.costo), 0) AS costo
            FROM ventas v
            LEFT JOIN (
                SELECT
                    venta_id,
                    SUM(cantidad * costo_unitario) AS costo
                FROM detalle_venta
                GROUP BY venta_id
            ) costos
                ON costos.venta_id = v.id
            WHERE v.cancelada = 0
              AND v.fecha = ?
            """,
            (closing_date,),
        )
        row = cursor.fetchone()

        cursor.execute(
            """
            SELECT
                COALESCE(dv.metodo_pago, v.metodo_pago) AS metodo_pago,
                COUNT(DISTINCT v.id) AS tickets,
                COALESCE(SUM(
                    CASE
                        WHEN dv.metodo_pago IS NULL AND v.subtotal > 0
                        THEN v.total * dv.subtotal / v.subtotal
                        ELSE dv.subtotal - dv.descuento
                    END
                ), 0) AS ventas,
                COALESCE(SUM(
                    COALESCE(
                        dv.monto_comision,
                        CASE WHEN v.subtotal > 0
                            THEN v.monto_comision * dv.subtotal / v.subtotal
                            ELSE 0 END
                    )
                ), 0) AS comisiones,
                COALESCE(SUM(
                    COALESCE(
                        dv.total_neto,
                        CASE
                            WHEN dv.metodo_pago IS NULL AND v.subtotal > 0
                            THEN v.total * dv.subtotal / v.subtotal
                            ELSE dv.subtotal - dv.descuento
                        END
                        - CASE WHEN v.subtotal > 0
                            THEN v.monto_comision * dv.subtotal / v.subtotal
                            ELSE 0 END
                    )
                ), 0) AS dinero_neto
            FROM detalle_venta dv
            INNER JOIN ventas v ON v.id = dv.venta_id
            WHERE v.cancelada = 0
              AND v.fecha = ?
            GROUP BY COALESCE(dv.metodo_pago, v.metodo_pago)
            """,
            (closing_date,),
        )
        payment_rows = {
            payment["metodo_pago"]: dict(payment)
            for payment in cursor.fetchall()
        }

        cursor.execute(
            """
            SELECT COALESCE(SUM(dv.cantidad), 0)
            FROM detalle_venta dv
            INNER JOIN ventas v
                ON v.id = dv.venta_id
            WHERE v.cancelada = 0
              AND v.fecha = ?
            """,
            (closing_date,),
        )
        products_sold = float(cursor.fetchone()[0])

        payments = []
        for method in self.PAYMENT_METHODS:
            values = payment_rows.get(method, {})
            payments.append(
                {
                    "metodo_pago": method,
                    "tickets": int(values.get("tickets", 0)),
                    "ventas": round(float(values.get("ventas", 0)), 2),
                    "comisiones": round(
                        float(values.get("comisiones", 0)), 2
                    ),
                    "dinero_neto": round(
                        float(values.get("dinero_neto", 0)), 2
                    ),
                }
            )

        sales = float(row["ventas"])
        commissions = float(row["comisiones"])
        net_money = float(row["dinero_neto"])
        cost = float(row["costo"])

        return {
            "fecha": closing_date,
            "formas_pago": payments,
            "total_ventas": round(sales, 2),
            "comisiones": round(commissions, 2),
            "dinero_neto": round(net_money, 2),
            "numero_tickets": int(row["tickets"]),
            "productos_vendidos": round(products_sold, 3),
            "costo": round(cost, 2),
            "utilidad": round(sales - commissions - cost, 2),
        }

    def calculate_cash_reconciliation(
        self,
        closing_date: str,
        counts: dict[tuple[str, float], int] | None = None,
    ) -> dict:
        """Compara el efectivo contado contra las ventas en efectivo."""

        counts = counts or {}
        rows = []
        total_counted = 0.0
        for kind, denomination in self.DENOMINATIONS:
            quantity = max(0, int(counts.get((kind, denomination), 0)))
            amount = round(denomination * quantity, 2)
            total_counted += amount
            rows.append({
                "tipo": kind,
                "denominacion": denomination,
                "cantidad": quantity,
                "importe": amount,
            })

        closing = self.get_daily_closing(closing_date)
        expected = next(
            payment["ventas"]
            for payment in closing["formas_pago"]
            if payment["metodo_pago"] == "Efectivo"
        )
        difference = round(total_counted - expected, 2)
        return {
            "fecha": closing_date,
            "denominaciones": rows,
            "efectivo_esperado": round(expected, 2),
            "efectivo_contado": round(total_counted, 2),
            "diferencia": difference,
            "estado": "CUADRA" if difference == 0 else (
                "SOBRANTE" if difference > 0 else "FALTANTE"
            ),
        }

    def export_excel(
        self,
        closing_date: str,
        destination: str | Path,
        cash_counts: dict[tuple[str, float], int] | None = None,
    ) -> Path:
        data = self.get_daily_closing(closing_date)
        cash = self.calculate_cash_reconciliation(closing_date, cash_counts)
        output_path = Path(destination)
        if output_path.suffix.lower() != ".xlsx":
            output_path = output_path.with_suffix(".xlsx")
        output_path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Corte diario"
        dark_fill = PatternFill("solid", fgColor="1F4E78")
        light_fill = PatternFill("solid", fgColor="D9EAF7")
        white_bold = Font(color="FFFFFF", bold=True)
        bold = Font(bold=True)
        money = '"$"#,##0.00'
        business = ConfigurationService(self.database).get_business_name()

        sheet.merge_cells("A1:E1")
        sheet["A1"] = f"{business} POS - Corte diario {closing_date}"
        sheet["A1"].fill = dark_fill
        sheet["A1"].font = Font(color="FFFFFF", bold=True, size=16)
        sheet["A1"].alignment = Alignment(horizontal="center")

        headers = (
            "Forma de pago", "Tickets", "Ventas", "Comisiones", "Dinero neto"
        )
        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(row=3, column=column, value=header)
            cell.fill = dark_fill
            cell.font = white_bold
            cell.alignment = Alignment(horizontal="center")

        for row_number, payment in enumerate(data["formas_pago"], start=4):
            values = (
                payment["metodo_pago"], payment["tickets"], payment["ventas"],
                payment["comisiones"], payment["dinero_neto"],
            )
            for column, value in enumerate(values, start=1):
                cell = sheet.cell(row=row_number, column=column, value=value)
                if column >= 3:
                    cell.number_format = money

        summary = (
            ("TOTAL VENTAS", data["total_ventas"], money),
            ("Comisiones Mercado Libre", data["comisiones"], money),
            ("DINERO NETO", data["dinero_neto"], money),
            ("Número de tickets", data["numero_tickets"], "#,##0"),
            ("Productos vendidos", data["productos_vendidos"], "#,##0.###"),
            ("Utilidad del día", data["utilidad"], money),
        )
        for row_number, (label, value, number_format) in enumerate(
            summary, start=9
        ):
            label_cell = sheet.cell(row=row_number, column=1, value=label)
            value_cell = sheet.cell(row=row_number, column=2, value=value)
            label_cell.fill = light_fill
            label_cell.font = bold
            value_cell.font = bold
            value_cell.number_format = number_format

        cash_start = 17
        sheet.merge_cells(
            start_row=cash_start, start_column=1,
            end_row=cash_start, end_column=4,
        )
        sheet.cell(cash_start, 1, "Arqueo manual de efectivo")
        sheet.cell(cash_start, 1).fill = dark_fill
        sheet.cell(cash_start, 1).font = white_bold
        cash_headers = ("Tipo", "Denominación", "Cantidad", "Importe")
        for column, header in enumerate(cash_headers, start=1):
            cell = sheet.cell(cash_start + 1, column, header)
            cell.fill = light_fill
            cell.font = bold
        row_number = cash_start + 2
        for denomination in cash["denominaciones"]:
            values = (
                denomination["tipo"], denomination["denominacion"],
                denomination["cantidad"], denomination["importe"],
            )
            for column, value in enumerate(values, start=1):
                sheet.cell(row_number, column, value)
            sheet.cell(row_number, 2).number_format = money
            sheet.cell(row_number, 4).number_format = money
            row_number += 1
        for label, value, format_code in (
            ("Efectivo esperado", cash["efectivo_esperado"], money),
            ("Efectivo contado", cash["efectivo_contado"], money),
            ("Diferencia", cash["diferencia"], money),
            ("Estado", cash["estado"], "@"),
        ):
            sheet.cell(row_number, 1, label).font = bold
            sheet.cell(row_number, 2, value).number_format = format_code
            row_number += 1

        widths = (28, 18, 18, 18, 18)
        for column, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + column)].width = width
        sheet.freeze_panes = "A4"
        workbook.save(output_path)
        return output_path

    def export_pdf(
        self,
        closing_date: str,
        destination: str | Path,
        cash_counts: dict[tuple[str, float], int] | None = None,
    ) -> Path:
        data = self.get_daily_closing(closing_date)
        cash = self.calculate_cash_reconciliation(closing_date, cash_counts)
        output_path = Path(destination)
        if output_path.suffix.lower() != ".pdf":
            output_path = output_path.with_suffix(".pdf")
        output_path.parent.mkdir(parents=True, exist_ok=True)

        business = ConfigurationService(self.database).get_business_name()
        styles = getSampleStyleSheet()
        story = []
        logo_path = Path(__file__).resolve().parents[1] / "assets" / "logo_ticket.png"
        if logo_path.is_file():
            logo = Image(str(logo_path), width=45 * mm, height=30 * mm)
            logo.hAlign = "CENTER"
            story.extend([logo, Spacer(1, 4 * mm)])

        story.append(Paragraph(f"<b>{business} POS</b>", styles["Title"]))
        story.append(
            Paragraph(f"Corte diario - {closing_date}", styles["Heading2"])
        )
        story.append(Spacer(1, 5 * mm))

        payment_data = [[
            "Forma de pago", "Tickets", "Ventas", "Comisiones", "Dinero neto"
        ]]
        for payment in data["formas_pago"]:
            payment_data.append([
                payment["metodo_pago"], str(payment["tickets"]),
                f"${payment['ventas']:,.2f}",
                f"${payment['comisiones']:,.2f}",
                f"${payment['dinero_neto']:,.2f}",
            ])
        payment_table = Table(payment_data, repeatRows=1)
        payment_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
            ("TOPPADDING", (0, 0), (-1, 0), 4),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 4),
        ]))
        story.extend([payment_table, Spacer(1, 4 * mm)])

        summary_data = [
            ["TOTAL VENTAS", f"${data['total_ventas']:,.2f}"],
            ["Comisiones Mercado Libre", f"${data['comisiones']:,.2f}"],
            ["DINERO NETO", f"${data['dinero_neto']:,.2f}"],
            ["Número de tickets", str(data["numero_tickets"])],
            ["Productos vendidos", f"{data['productos_vendidos']:g}"],
            ["Utilidad del día", f"${data['utilidad']:,.2f}"],
        ]
        summary_table = Table(summary_data, colWidths=[70 * mm, 45 * mm])
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#D9EAF7")),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("PADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 4 * mm))
        story.append(Paragraph("<b>Arqueo manual de efectivo</b>", styles["Heading2"]))
        cash_data = [["Tipo", "Denominación", "Cantidad", "Importe"]]
        for denomination in cash["denominaciones"]:
            cash_data.append([
                denomination["tipo"],
                f"${denomination['denominacion']:,.2f}",
                str(denomination["cantidad"]),
                f"${denomination['importe']:,.2f}",
            ])
        cash_table = Table(cash_data, repeatRows=1)
        cash_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.extend([cash_table, Spacer(1, 4 * mm)])
        reconciliation = Table([
            ["Efectivo esperado", f"${cash['efectivo_esperado']:,.2f}"],
            ["Efectivo contado", f"${cash['efectivo_contado']:,.2f}"],
            ["Diferencia", f"${cash['diferencia']:,.2f}"],
            ["Estado", cash["estado"]],
        ], colWidths=[70 * mm, 45 * mm])
        reconciliation.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#D9EAF7")),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(reconciliation)

        document = SimpleDocTemplate(
            str(output_path), pagesize=letter,
            rightMargin=18 * mm, leftMargin=18 * mm,
            topMargin=8 * mm, bottomMargin=8 * mm,
            title=f"Corte diario {closing_date}",
        )
        document.build(story)
        return output_path
